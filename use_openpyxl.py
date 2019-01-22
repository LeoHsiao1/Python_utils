from openpyxl import Workbook
from openpyxl import load_workbook  # 从openpyxl模块导入Workbook类，用于创建和写入表格

# openpyxl是一个用于读写 Excel 2010 xlsx/xlsm 文件的库，在终端输入命令pip install openpyxl即可安装。
# 一般操作流程为：先打开Workbook（即工作簿），再定位Worksheet（即工作表），然后操作Cell（即单元格）。


# 基本操作如下：
def sample():
    wb = Workbook()                     # 创建一个工作簿对象
    print(wb.sheetnames)                # 列出所有工作表的名字，新创建的工作簿只有一个默认工作表
    ws1 = wb.create_sheet("sheet1", 0)  # 在第一个位置插入新工作表。如果不指定位置就默认在最后插入
    ws1.title = "Sheet1"                # 设置工作表的名字

    # 使用单元格的坐标作为键值，访问指定单元格
    ws1['A1'] = 2
    print(ws1['A1'].value)  # 通过value属性获得该单元格的值

    # 通过cell(row,column,value=None)方法访问指定单元格
    ws1.cell(1, 2, 'Hi')
    print(ws1.cell(1, 2).value)

    # 通过append()方法在工作表后附加一行（工作表对象没有插入的方法）
    ws1.append(["Hello", "World", "!"])

    # 把工作簿保存成一个xlsx文件
    wb.save(r"C:\Users\Will\Desktop\1.xlsx")

    # 读取一个工作簿
    wb = load_workbook(r"C:\Users\Will\Desktop\1.xlsx")

    ws1 = wb[wb.sheetnames[0]]                # 用工作表的名字作为工作簿的键值，可得到该工作表对象的引用
    print(ws1.max_row)                        # 获得工作表的最大行数
    print(ws1.max_column)                     # 获得工作表的最大列数
    for row in ws1.values:                    # 使用工作表的values属性遍历所有单元格的值，如果值为空就返回None
        print(row)

    wb.close()                                # 关闭对该文件的占用，但该文件的内容仍然保留在内存中可以读取
    # 保存工作簿时，如果存在同名文件，会直接覆盖而不会提醒。如果其它程序正在使用该文件就无法保存。


sample()


def save_xlsx():
    pass
