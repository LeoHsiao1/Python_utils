"""
实现一些与终端交互、读写文件的功能。
  class `Inputs`
  def `print_text`
  def `read_csv`
  def `write_csv`
  def `read_xlsx`
  def `write_xlsx`
"""

import os
import sys
import time


class Inputs:
    """
    提供一些从终端获取用户输入的静态方法。
      - 可在外部创建该类的变量，暂存一些输入的数据。
    """

    @staticmethod
    def input_int(tip_str='', retry=True):
        while retry:
            try:
                return int(input(tip_str))
            except:
                print("输入的不是整数！")

    @staticmethod
    def input_positive_int(tip_str='', retry=True):
        while retry:
            num = Inputs.input_int(tip_str, retry)
            if num > 0:
                return num
            else:
                print("输入的不是正整数！")

    @staticmethod
    def input_real_num(tip_str='', retry=True):
        while retry:
            try:
                return float(input(tip_str))
            except:
                print("输入的不是整数或浮点数！")

    @staticmethod
    def input_path(tip_str='', retry=True):
        while retry:
            path = input(tip_str)
            if os.path.isdir(path):
                return path
            else:
                print("输入的不是有效目录！")


def print_text(text, delay=0):
    """ 在DOS窗口中显示文本text，显示每个字符的间隔时长为delay """
    for line in text:
        for word in line:
            print(word, end='')  # 逐个字显示
            time.sleep(delay)
            # 每打印一个字符就刷新一次stdout，
            # 否则缓存区累积了一行才会刷新stdout
            sys.stdout.flush()


def read_csv(filename, *args, **kwargs):
    """
    从一个csv文件中读取数据，并将每行数据转换成list格式。
      - 基于csv模块。
      - 该函数的参数列表与open()相同。
    """
    import csv
    result = []
    with open(filename, *args, **kwargs) as f:
        csv_reader = csv.reader(f)   # 在一个文件流f上创建csv阅读器
        try:
            for line in csv_reader:  # 迭代csv_reader的内容
                result.append(line)
        except csv.Error as e:       # 捕捉读取csv文件时的异常
            raise OSError("A reading error in file {}, line {}:\n{}"
                          .format(filename, csv_reader.line_num, e))
    return result


def write_csv(data, filename, mode='w', newline='', **kwargs):
    """ 
    将数据转换成csv格式，再保存到指定文件中。
      - 基于csv模块。
    """
    import csv
    with open(filename, mode, newline=newline, **kwargs) as f:
        csv_writer = csv.writer(f)  # 在文件流f上创建一个csv写入器
        csv_writer.writerows(data)  # 写入多行


def read_xlsx(filename, read_only=True, *args, **kwargs):
    """
    读取一个xlsx表格中的全部数据，保存为一个字典返回。
      - 该字典的key为xlsx中一个sheet的名字，value为该sheet的所有行组成的list。
      - 该函数的参数与 openpyxl.load_workbook() 相同。
      - read_only=True 表示以只读模式打开，读取速度更快。
    """
    from openpyxl import load_workbook
    wb = load_workbook(filename, read_only, *args, **kwargs)
    data_dict = {}
    # 遍历xlsx中的每个sheet，遍历每个sheet中的每行数据，保存为字典类型
    for name in wb.sheetnames:
        data_dict[name] = [row for row in wb[name].values]
    wb.close()
    return data_dict


def write_xlsx(data_dict, filename, write_only=True):
    """
    把一个字典写入xlsx表格。
      - 如果输入的data不是字典类型，会先转换成 key="Sheet1" 的字典。
      - 该字典的key为xlsx中一个sheet的名字，value为该sheet的所有行组成的list。
      - write_only=True 表示以只写模式打开，读取速度更快。
    """
    from openpyxl import Workbook

    if not isinstance(data_dict, dict):
        data_dict = {"Sheet1": data_dict}

    wb = Workbook(data_dict, write_only)
    # 遍历data_dict中的每个value，遍历每个value的每行数据，保存为xlsx表格
    for k, v in data_dict.items():
        ws = wb.create_sheet(k)
        for row in v:
            ws.append(row)
    wb.save(filename)
    wb.close()